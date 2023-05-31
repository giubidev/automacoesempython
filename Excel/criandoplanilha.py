import openpyxl

# Criar uma planilha
book = openpyxl.Workbook()

# Visualizar páginas existentes
print(book.sheetnames)

# Criar uma página
book.create_sheet('Frutas')

# Selecionar uma página
frutaspg = book['Frutas']

# Adicionar dados na página
frutaspg.append(['Banana', '5', 'R$3,90'])
frutaspg.append(['Fruta 2', '2', 'R$15,90'])
frutaspg.append(['Fruta 3', '10', 'R$31,90'])
frutaspg.append(['Fruta 4', '3', 'R$9,90'])

# Salvar a planilha
book.save('Planilha de Compras.xlsx')


